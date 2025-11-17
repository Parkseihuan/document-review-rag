"""XLSX file parser"""
from typing import Optional
import openpyxl
from openpyxl.utils import get_column_letter


class XLSXParser:
    """Parse XLSX files to markdown"""

    @staticmethod
    def parse(file_path: str) -> Optional[str]:
        """
        Parse XLSX file and extract data as markdown tables

        Args:
            file_path: Path to XLSX file

        Returns:
            Text content in markdown format or None if parsing fails
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            markdown_parts = []

            # Add title
            filename = file_path.split('/')[-1].replace('.xlsx', '')
            markdown_parts.append(f"# {filename}\n")

            # Process each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Add sheet name as heading
                markdown_parts.append(f"\n## {sheet_name}\n")

                # Get all rows
                rows = list(sheet.iter_rows(values_only=True))

                if not rows:
                    markdown_parts.append("*Empty sheet*\n")
                    continue

                # Find the actual data range (skip empty rows/columns)
                non_empty_rows = []
                for row in rows:
                    # Check if row has any non-empty cells
                    if any(cell is not None and str(cell).strip() for cell in row):
                        non_empty_rows.append(row)

                if not non_empty_rows:
                    markdown_parts.append("*Empty sheet*\n")
                    continue

                # Find max column count
                max_cols = max(len(row) for row in non_empty_rows)

                # Create markdown table
                table_rows = []
                for i, row in enumerate(non_empty_rows):
                    # Pad row to max_cols length
                    padded_row = list(row) + [None] * (max_cols - len(row))
                    # Convert cells to strings
                    cells = [str(cell) if cell is not None else '' for cell in padded_row]
                    table_rows.append('| ' + ' | '.join(cells) + ' |')

                    # Add header separator after first row
                    if i == 0:
                        table_rows.append('| ' + ' | '.join(['---'] * max_cols) + ' |')

                markdown_parts.append('\n'.join(table_rows))
                markdown_parts.append('\n')

            workbook.close()
            return '\n'.join(markdown_parts)

        except Exception as e:
            print(f"Error parsing XLSX {file_path}: {e}")
            return None

    @staticmethod
    def to_markdown(file_path: str, output_path: str) -> bool:
        """
        Parse XLSX and save as markdown file

        Args:
            file_path: Path to XLSX file
            output_path: Path to save markdown file

        Returns:
            True if successful, False otherwise
        """
        content = XLSXParser.parse(file_path)
        if content:
            try:
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                return True
            except Exception as e:
                print(f"Error writing markdown file: {e}")
                return False
        return False
